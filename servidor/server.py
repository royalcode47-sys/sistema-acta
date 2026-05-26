"""
AIR-E · Servidor de Actas
Llena la plantilla MONITOR.xlsx con los datos del formulario y la devuelve para descarga.
Sirve también el frontend estático (index.html, css/, js/, assets/).

Requisitos:
    pip install flask flask-cors openpyxl gunicorn

Uso local:
    python servidor/server.py
    El servidor corre en http://localhost:5000

Despliegue en Render:
    Build Command : pip install -r requirements.txt
    Start Command : gunicorn --chdir servidor server:app --bind 0.0.0.0:$PORT
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
import io
import os
from datetime import datetime

# ── Rutas ────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))   # .../servidor/
FRONTEND_DIR = os.path.join(BASE_DIR, "..")                  # raíz del proyecto
PLANTILLA    = os.path.join(BASE_DIR, "MONITOR_plantilla.xlsx")

# ── App Flask ─────────────────────────────────────────────────────────────────
app = Flask(
    __name__,
    static_folder=FRONTEND_DIR,   # sirve css/, js/, assets/
    static_url_path=""
)
CORS(app)


# ── Ruta principal → index.html ───────────────────────────────────────────────
@app.route("/")
def index():
    return send_file(os.path.join(FRONTEND_DIR, "index.html"))


# ── Ping ─────────────────────────────────────────────────────────────────────
@app.route("/ping", methods=["GET"])
def ping():
    return jsonify({"status": "ok", "message": "Servidor AIR-E activo"})


# ── Generar Acta ──────────────────────────────────────────────────────────────
def generar_acta(data: dict) -> io.BytesIO:
    """
    Abre la plantilla, llena las celdas con los datos recibidos
    y retorna el archivo en memoria listo para descarga.

    Mapa de celdas (hoja 'Entrega_Dev disp'):
    ─────────────────────────────────────────
    QUIEN ENTREGA (Activo entregado por)
      B10 → nombre_entrega
      B11 → cedula_entrega
      B12 → cargo_entrega
      B13 → empresa_entrega

    QUIEN RECIBE (Activo entregado a)
      E10 → nombre_recibe
      E11 → cedula_recibe
      E12 → cargo_recibe
      E13 → empresa_recibe

    UBICACIÓN
      H10 → ciudad
      H11 → sede
      H12 → piso

    TIPO DE GESTIÓN (marcar con X)
      E7  → 'X' si tipo == 'asignacion' o 'cambio'
      I7  → 'X' si tipo == 'devolucion' o 'cambio'

    TICKET
      B7  → numero_ticket

    EQUIPOS (filas 17-24, se llenan dinámicamente)
      Columnas: A=tipo, B=placa, C=serial, D=modelo, E=marca, F=observacion
    """

    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb["Entrega_Dev disp"]

    tipo = data.get("tipo", "asignacion")  # asignacion | cambio | devolucion

    # ── Ticket ──────────────────────────────────────────────
    ws["B7"] = data.get("ticket", "")

    # ── Tipo de gestión ─────────────────────────────────────
    if tipo == "asignacion":
        ws["E7"] = "X"
        ws["I7"] = ""
    elif tipo == "cambio":
        ws["E7"] = "X"   # Entrega de equipo nuevo
        ws["I7"] = "X"   # Devolución del equipo anterior
    else:  # devolucion
        ws["E7"] = ""
        ws["I7"] = "X"

    # ── Quien entrega (TI / empresa) ────────────────────────
    ws["B10"] = data.get("nombre_entrega", "")
    ws["B11"] = data.get("cedula_entrega", "")
    ws["B12"] = data.get("cargo_entrega", "")
    ws["B13"] = data.get("empresa_entrega", "AIR-E")

    # ── Quien recibe (empleado seleccionado) ────────────────
    ws["E10"] = data.get("nombre_recibe", "")
    ws["E11"] = data.get("cedula_recibe", "")
    ws["E12"] = data.get("cargo_recibe", "")
    ws["E13"] = data.get("empresa_recibe", "AIR-E")

    # ── Ubicación ───────────────────────────────────────────
    ws["H10"] = data.get("ciudad", "BARRANQUILLA")
    ws["H11"] = data.get("sede", "")
    ws["H12"] = data.get("piso", "")

    # ── Equipos ─────────────────────────────────────────────
    for fila in range(17, 25):
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{fila}"] = None

    equipos = data.get("equipos", [])
    for i, equipo in enumerate(equipos[:8]):
        fila = 17 + i
        ws[f"A{fila}"] = equipo.get("tipo_activo", "")
        ws[f"B{fila}"] = equipo.get("placa", "")
        ws[f"C{fila}"] = equipo.get("serial", "")
        ws[f"D{fila}"] = equipo.get("modelo", "")
        ws[f"E{fila}"] = equipo.get("marca", "")
        ws[f"F{fila}"] = equipo.get("observacion", "")

    # ── Características técnicas (si es computador) ─────────
    specs = data.get("specs", {})
    if specs:
        ws["C27"] = specs.get("nombre_pc", "")
        ws["C28"] = specs.get("procesador", "")
        ws["C29"] = specs.get("ram", "")
        ws["C30"] = specs.get("disco", "")
        ws["C31"] = specs.get("opticos", "")
        ws["C32"] = specs.get("monitor", "")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.route("/generar-acta", methods=["POST"])
def endpoint_generar_acta():
    try:
        data = request.get_json(force=True)
        if not data:
            return jsonify({"error": "No se recibieron datos"}), 400

        buffer = generar_acta(data)

        nombre_empleado = data.get("nombre_recibe", "empleado").replace(" ", "_")
        fecha = datetime.now().strftime("%Y%m%d")
        tipo_label = {
            "asignacion": "Asignacion",
            "cambio": "Cambio",
            "devolucion": "Devolucion"
        }.get(data.get("tipo", "asignacion"), "Acta")

        filename = f"Acta_{tipo_label}_{nombre_empleado}_{fecha}.xlsx"

        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Arranque ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") != "production"

    if not os.path.exists(PLANTILLA):
        print(f"ERROR: No se encontró la plantilla en {PLANTILLA}")
    else:
        print("✅ Plantilla encontrada")
        print(f"🚀 Servidor AIR-E corriendo en http://localhost:{port}")
        app.run(debug=debug, port=port, host="0.0.0.0")